from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from .models import Finance
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@login_required
def finance_list(request):
    from erp_config.models import Building, Category
    # period filter
    period   = request.GET.get('period', 'all')
    finances = Finance.objects.all().order_by('-expense_date')
    today    = date.today()

    if period == 'today':
        finances = finances.filter(expense_date=today)
    elif period == 'week':
        finances = finances.filter(expense_date__week=today.isocalendar()[1],
                                   expense_date__year=today.year)
    elif period == 'month':
        finances = finances.filter(expense_date__month=today.month,
                                   expense_date__year=today.year)
    elif period == 'year':
        finances = finances.filter(expense_date__year=today.year)
    elif period == 'custom':
        start = request.GET.get('start')
        end   = request.GET.get('end')
        if start and end:
            finances = finances.filter(expense_date__range=[start, end])

    total     = sum(f.amount for f in finances)

    paginator  = Paginator(finances, 10)
    page       = request.GET.get('page')
    finances   = paginator.get_page(page)
    buildings  = Building.objects.all()
    categories = Category.objects.all()

    return render(request, 'finance/list.html', {
        'finances':   finances,
        'total':      total,
        'period':     period,
        'buildings':  buildings,
        'categories': categories,
    })

@login_required
def finance_save(request):
    if request.method == 'POST':
        Finance.objects.create(
            expense_date = request.POST['expense_date'],
            nature       = request.POST['nature'],
            building     = request.POST['building'],
            amount       = request.POST['amount'],
            remarks      = request.POST.get('remarks', ''),
            person       = request.POST['person'],
        )
        messages.success(request, 'Expense saved successfully!')
    return redirect('finance_list')

@login_required
def finance_delete(request, pk):
    finance = get_object_or_404(Finance, pk=pk)
    finance.delete()
    messages.success(request, 'Record deleted!')
    return redirect('finance_list')

@login_required
def finance_update(request, pk):
    finance = get_object_or_404(Finance, pk=pk)
    if request.method == 'POST':
        finance.expense_date = request.POST['expense_date']
        finance.nature       = request.POST['nature']
        finance.building     = request.POST['building']
        finance.amount       = request.POST['amount']
        finance.remarks      = request.POST.get('remarks', '')
        finance.person       = request.POST['person']
        finance.save()
        messages.success(request, 'Expense record updated!')
        return redirect('finance_list')
    from erp_config.models import Building, Category
    buildings  = Building.objects.filter(type='Growing')
    categories = Category.objects.all()
    return render(request, 'finance/edit.html', {
        'finance':    finance,
        'buildings':  buildings,
        'categories': categories,
    })

@login_required
def export_finance(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Finance'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='D4880A')

    headers = ['Date', 'Nature', 'Building', 'Amount', 'Person', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, f in enumerate(Finance.objects.all().order_by('-expense_date'), 2):
        ws.cell(row=row, column=1, value=str(f.expense_date))
        ws.cell(row=row, column=2, value=f.nature)
        ws.cell(row=row, column=3, value=f.building)
        ws.cell(row=row, column=4, value=float(f.amount))
        ws.cell(row=row, column=5, value=f.person)
        ws.cell(row=row, column=6, value=f.remarks)

    for col, width in enumerate([14, 16, 18, 12, 16, 25], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="finance_{date.today()}.xlsx"'
    wb.save(response)
    return response