from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import F
from .models import Consumption
from stock.models import Stock
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from datetime import date

# ── FIFO deduction helper ─────────────────────────────────
def fifo_deduct(item_id, quantity, consumption=None):
    from .models import ConsumptionBatchDeduction
    batches = Stock.objects.filter(
        item_id=item_id,
        quantity__gt=0
    ).order_by(F('expiry_date').asc(nulls_last=True), 'date')

    if not batches.exists():
        return False, 0, ''

    item_name       = batches.first().name
    total_available = sum(int(b.quantity) for b in batches)

    if quantity > total_available:
        return False, total_available, item_name

    remaining = quantity
    for batch in batches:
        if remaining <= 0:
            break
        batch_qty = int(batch.quantity)
        if batch_qty >= remaining:
            deducted       = remaining
            batch.quantity = batch_qty - remaining
            batch.save()
            remaining      = 0
        else:
            deducted       = batch_qty
            remaining     -= batch_qty
            batch.quantity = 0
            batch.save()

        if consumption:
            ConsumptionBatchDeduction.objects.create(
                consumption       = consumption,
                stock_batch       = batch,
                quantity_deducted = deducted
            )

    total_left = sum(int(q) for q in Stock.objects.filter(item_id=item_id).values_list('quantity', flat=True))
    return True, total_left, item_name

def fifo_restore(consumption):
    from .models import ConsumptionBatchDeduction
    deductions = ConsumptionBatchDeduction.objects.filter(consumption=consumption)
    if deductions.exists():
        for deduction in deductions:
            if deduction.stock_batch:
                deduction.stock_batch.quantity = int(deduction.stock_batch.quantity) + int(deduction.quantity_deducted)
                deduction.stock_batch.save()
        deductions.delete()
    else:
        batch = Stock.objects.filter(item_id=consumption.item_id).order_by(
            F('expiry_date').desc(nulls_last=True), '-date'
        ).first()
        if batch:
            batch.quantity = int(batch.quantity) + int(consumption.quantity)
            batch.save()

@login_required
def consumption_list(request):
    from erp_config.models import Category, Unit, Building
    growing_buildings = Building.objects.filter(type='Growing').values_list('name', flat=True)
    consumption_all = Consumption.objects.filter(
                        growing_house__in=growing_buildings
                      ).order_by('-date_consumed')
    paginator       = Paginator(consumption_all, 10)
    page            = request.GET.get('page')
    consumptions    = paginator.get_page(page)
    categories      = Category.objects.all()
    units           = Unit.objects.all()
    buildings       = Building.objects.filter(type='Growing')
    stock_items     = Stock.objects.filter(quantity__gt=0).order_by('item_id').values(
                        'item_id', 'name', 'category', 'unit'
                      ).distinct()
    from master_data.models import Material
    materials = Material.objects.all()
    return render(request, 'consumption/list.html', {
        'consumptions': consumptions,
        'categories':   categories,
        'units':        units,
        'buildings':    buildings,
        'stock_items':  stock_items,
        'materials':    materials,
    })

@login_required
def consumption_save(request):
    if request.method == 'POST':
        from .models import ConsumptionBatchDeduction
        item_id       = request.POST['item_id']
        house         = request.POST['growing_house']
        category      = request.POST['category']
        quantity      = float(request.POST['quantity'])
        unit          = request.POST['unit']
        remarks       = request.POST.get('remarks', '')
        recorded      = request.POST['recorded_by']
        date_consumed = request.POST['date_consumed']

        from master_data.models import Material
        from decimal import Decimal

        # UoM conversion
        consumed_unit     = request.POST.get('consumed_unit', unit)
        original_quantity = quantity
        original_unit     = consumed_unit
        try:
            material = Material.objects.get(item_id=item_id)
            if material.stock_unit and material.conversion_factor and consumed_unit == material.stock_unit:
                quantity = float(Decimal(str(quantity)) * Decimal(str(material.conversion_factor)))
                unit     = material.base_unit or unit
        except Material.DoesNotExist:
            pass

        batches = Stock.objects.filter(
            item_id=item_id, quantity__gt=0
        ).order_by(F('expiry_date').asc(nulls_last=True), 'date')

        if not batches.exists():
            messages.error(request, 'Item not found in stock!')
            return redirect('consumption_list')

        item_name       = batches.first().name
        total_available = sum(int(b.quantity) for b in batches)

        if quantity > total_available:
            messages.error(request, f'Not enough stock! Available: {total_available} {unit}')
            return redirect('consumption_list')

        consumption = Consumption.objects.create(
            growing_house = house,
            category      = category,
            item_id       = item_id,
            item_name     = item_name,
            quantity      = quantity,
            unit          = unit,
            remarks       = remarks,
            recorded_by   = recorded,
            date_consumed = date_consumed,
            original_quantity = original_quantity,
            original_unit     = original_unit,
        )

        remaining = quantity
        for batch in batches:
            if remaining <= 0:
                break
            batch_qty = int(batch.quantity)
            if batch_qty >= remaining:
                deducted       = remaining
                batch.quantity = batch_qty - remaining
                batch.save()
                remaining      = 0
            else:
                deducted       = batch_qty
                remaining     -= batch_qty
                batch.quantity = 0
                batch.save()

            ConsumptionBatchDeduction.objects.create(
                consumption       = consumption,
                stock_batch       = batch,
                quantity_deducted = deducted
            )

        total_left = sum(int(q) for q in Stock.objects.filter(
            item_id=item_id).values_list('quantity', flat=True))
        messages.success(request, f'Consumption saved! Remaining {item_name}: {total_left}')
    return redirect('consumption_list')

@login_required
def consumption_delete(request, pk):
    consumption = get_object_or_404(Consumption, pk=pk)
    fifo_restore(consumption)
    consumption.delete()
    messages.success(request, 'Record deleted and stock restored!')
    return redirect('consumption_list')

@login_required
def get_items(request):
    from django.db.models import Sum
    from master_data.models import Material
    category = request.GET.get('category', '')
    building = request.GET.get('building', '')

    filters = {'quantity__gt': 0}
    if category:
        filters['category'] = category
    if building:
        filters['growing_house'] = building

    items = Stock.objects.filter(**filters).values(
        'item_id', 'name', 'unit'
    ).annotate(
        quantity=Sum('quantity')
    ).order_by('name')

    result = []
    for item in items:
        material = Material.objects.filter(item_id=item['item_id']).first()
        result.append({
            'item_id':  item['item_id'],
            'name':     item['name'],
            'quantity': item['quantity'],
            'unit':     material.unit if material else item['unit'],
        })
    return JsonResponse({'items': result})

@login_required
def consumption_update(request, pk):
    from erp_config.models import Category, Unit, Building
    consumption = get_object_or_404(Consumption, pk=pk)
    if request.method == 'POST':
        # restore old stock first
        fifo_restore(consumption)

        # update fields
        consumption.growing_house = request.POST['growing_house']
        consumption.category      = request.POST['category']
        consumption.item_id       = request.POST['item_id']
        consumption.item_name     = request.POST['item_name']
        consumption.quantity      = float(request.POST['quantity'])
        consumption.unit          = request.POST['unit']
        consumption.remarks       = request.POST.get('remarks', '')
        consumption.recorded_by   = request.POST['recorded_by']
        consumption.date_consumed = request.POST['date_consumed']
        consumption.save()

        # re-deduct stock with FIFO
        success, remaining, item_name = fifo_deduct(
            consumption.item_id,
            consumption.quantity,
            consumption
        )
        if not success:
            messages.error(request, f'Not enough stock! Available: {remaining}')
        else:
            messages.success(request, f'Record updated! Remaining {item_name}: {remaining}')
        return redirect('consumption_list')

    categories = Category.objects.all()
    units      = Unit.objects.all()
    buildings  = Building.objects.filter(type='Growing')
    return render(request, 'consumption/edit.html', {
        'consumption': consumption,
        'categories':  categories,
        'units':       units,
        'buildings':   buildings,
    })

@login_required
def export_consumption(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Consumption'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='17A98A')

    headers = ['Date', 'Growing House', 'Category', 'Item ID',
               'Item Name', 'Quantity', 'Unit', 'Recorded By', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    from erp_config.models import Building
    growing_buildings = Building.objects.filter(type='Growing').values_list('name', flat=True)
    for row, c in enumerate(Consumption.objects.filter(
            growing_house__in=growing_buildings).order_by('-date_consumed'), 2):
        ws.cell(row=row, column=1, value=str(c.date_consumed))
        ws.cell(row=row, column=2, value=c.growing_house)
        ws.cell(row=row, column=3, value=c.category)
        ws.cell(row=row, column=4, value=c.item_id)
        ws.cell(row=row, column=5, value=c.item_name)
        ws.cell(row=row, column=6, value=float(c.quantity))
        ws.cell(row=row, column=7, value=c.unit)
        ws.cell(row=row, column=8, value=c.recorded_by)
        ws.cell(row=row, column=9, value=c.remarks)

    for col, width in enumerate([14, 18, 15, 12, 20, 10, 10, 16, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="consumption_{date.today()}.xlsx"'
    wb.save(response)
    return response

# ── Laying Consumption ────────────────────────────────────

@login_required
def laying_consumption_list(request):
    from erp_config.models import Category, Unit, Building
    laying_buildings = Building.objects.filter(type='Laying').values_list('name', flat=True)
    consumption_all  = Consumption.objects.filter(
                        growing_house__in=laying_buildings
                       ).order_by('-date_consumed')
    paginator    = Paginator(consumption_all, 10)
    page         = request.GET.get('page')
    consumptions = paginator.get_page(page)
    categories   = Category.objects.all()
    units        = Unit.objects.all()
    buildings    = Building.objects.filter(type='Laying')
    stock_items  = Stock.objects.filter(quantity__gt=0).order_by('item_id').values(
                    'item_id', 'name', 'category', 'unit'
                   ).distinct()
    return render(request, 'consumption/laying_list.html', {
        'consumptions': consumptions,
        'categories':   categories,
        'units':        units,
        'buildings':    buildings,
        'stock_items':  stock_items,
    })

@login_required
def laying_consumption_save(request):
    if request.method == 'POST':
        from .models import ConsumptionBatchDeduction
        item_id       = request.POST['item_id']
        house         = request.POST['growing_house']
        category      = request.POST['category']
        quantity      = float(request.POST['quantity'])
        unit          = request.POST['unit']
        remarks       = request.POST.get('remarks', '')
        recorded      = request.POST['recorded_by']
        date_consumed = request.POST['date_consumed']

        from master_data.models import Material
        from decimal import Decimal

        # UoM conversion
        consumed_unit = request.POST.get('consumed_unit', unit)
        try:
            material = Material.objects.get(item_id=item_id)
            if material.stock_unit and material.conversion_factor and consumed_unit == material.stock_unit:
                # staff entered in stock units — convert to base units
                quantity = float(Decimal(str(quantity)) * Decimal(str(material.conversion_factor)))
                unit     = material.base_unit or unit
        except Material.DoesNotExist:
            pass

        batches = Stock.objects.filter(
            item_id=item_id, quantity__gt=0
        ).order_by(F('expiry_date').asc(nulls_last=True), 'date')

        if not batches.exists():
            messages.error(request, 'Item not found in stock!')
            return redirect('laying_consumption_list')

        item_name       = batches.first().name
        total_available = sum(int(b.quantity) for b in batches)

        if quantity > total_available:
            messages.error(request, f'Not enough stock! Available: {total_available} {unit}')
            return redirect('laying_consumption_list')

        consumption = Consumption.objects.create(
            growing_house = house,
            category      = category,
            item_id       = item_id,
            item_name     = item_name,
            quantity      = quantity,
            unit          = unit,
            remarks       = remarks,
            recorded_by   = recorded,
            date_consumed = date_consumed,
            original_quantity = original_quantity,
            original_unit     = original_unit,
        )

        remaining = quantity
        for batch in batches:
            if remaining <= 0:
                break
            batch_qty = int(batch.quantity)
            if batch_qty >= remaining:
                deducted       = remaining
                batch.quantity = batch_qty - remaining
                batch.save()
                remaining      = 0
            else:
                deducted       = batch_qty
                remaining     -= batch_qty
                batch.quantity = 0
                batch.save()

            ConsumptionBatchDeduction.objects.create(
                consumption       = consumption,
                stock_batch       = batch,
                quantity_deducted = deducted
            )

        total_left = sum(int(q) for q in Stock.objects.filter(
            item_id=item_id).values_list('quantity', flat=True))
        messages.success(request, f'Consumption saved! Remaining {item_name}: {total_left}')
    return redirect('laying_consumption_list')

@login_required
def export_laying_consumption(request):
    from erp_config.models import Building
    laying_buildings = Building.objects.filter(type='Laying').values_list('name', flat=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laying Consumption'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='D4880A')

    headers = ['Date', 'Building', 'Category', 'Item ID',
               'Item Name', 'Quantity', 'Unit', 'Recorded By', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, c in enumerate(Consumption.objects.filter(
            growing_house__in=laying_buildings).order_by('-date_consumed'), 2):
        ws.cell(row=row, column=1, value=str(c.date_consumed))
        ws.cell(row=row, column=2, value=c.growing_house)
        ws.cell(row=row, column=3, value=c.category)
        ws.cell(row=row, column=4, value=c.item_id)
        ws.cell(row=row, column=5, value=c.item_name)
        ws.cell(row=row, column=6, value=float(c.quantity))
        ws.cell(row=row, column=7, value=c.unit)
        ws.cell(row=row, column=8, value=c.recorded_by)
        ws.cell(row=row, column=9, value=c.remarks)

    for col, width in enumerate([14, 18, 15, 12, 20, 10, 10, 16, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="laying_consumption_{date.today()}.xlsx"'
    wb.save(response)
    return response

@login_required
def laying_consumption_delete(request, pk):
    consumption = get_object_or_404(Consumption, pk=pk)
    fifo_restore(consumption)
    consumption.delete()
    messages.success(request, 'Record deleted and stock restored!')
    return redirect('laying_consumption_list')

@login_required
def laying_consumption_update(request, pk):
    from erp_config.models import Category, Unit, Building
    consumption = get_object_or_404(Consumption, pk=pk)
    if request.method == 'POST':
        # restore old stock first
        fifo_restore(consumption)

        # update fields
        consumption.growing_house = request.POST['growing_house']
        consumption.category      = request.POST['category']
        consumption.item_id       = request.POST['item_id']
        consumption.item_name     = request.POST['item_name']
        consumption.quantity      = float(request.POST['quantity'])
        consumption.unit          = request.POST['unit']
        consumption.remarks       = request.POST.get('remarks', '')
        consumption.recorded_by   = request.POST['recorded_by']
        consumption.date_consumed = request.POST['date_consumed']
        consumption.save()

        # re-deduct stock with FIFO
        success, remaining, item_name = fifo_deduct(
            consumption.item_id,
            consumption.quantity,
            consumption
        )
        if not success:
            messages.error(request, f'Not enough stock! Available: {remaining}')
        else:
            messages.success(request, f'Record updated! Remaining {item_name}: {remaining}')
        return redirect('laying_consumption_list')

    categories = Category.objects.all()
    units      = Unit.objects.all()
    buildings  = Building.objects.filter(type='Laying')
    return render(request, 'consumption/laying_edit.html', {
        'consumption': consumption,
        'categories':  categories,
        'units':       units,
        'buildings':   buildings,
    })