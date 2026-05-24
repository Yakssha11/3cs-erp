from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum
from .models import GrowingSale, LayingSale
from flock.models import Flock
from laying_flock.models import LayingFlock
from master_data.models import Customer
from erp_config.models import EggPriceConfig, Building
from finance.models import Finance
from laying_finance.models import LayingFinance
from consumption.models import Consumption
from datetime import date

@login_required
def growing_sales_list(request):
    sales    = GrowingSale.objects.all().order_by('-sale_date')
    flocks   = Flock.objects.filter(status='Active')
    customers = Customer.objects.all()
    total    = float(sales.aggregate(Sum('total_revenue'))['total_revenue__sum'] or 0)
    return render(request, 'sales/growing.html', {
        'sales':     sales,
        'flocks':    flocks,
        'customers': customers,
        'total':     total,
    })

@login_required
def growing_sale_save(request):
    if request.method == 'POST':
        flock_id      = request.POST['flock_id']
        customer_id   = request.POST.get('customer_id') or None
        quantity      = int(request.POST['quantity'])
        price_per_head = float(request.POST['price_per_head'])
        total_revenue  = round(quantity * price_per_head, 2)
        sale_date      = request.POST['sale_date']
        recorded_by    = request.POST['recorded_by']
        remarks        = request.POST.get('remarks', '')

        flock = get_object_or_404(Flock, pk=flock_id)

        GrowingSale.objects.create(
            flock_id       = flock_id,
            customer_id    = customer_id,
            quantity       = quantity,
            price_per_head = price_per_head,
            total_revenue  = total_revenue,
            sale_date      = sale_date,
            recorded_by    = recorded_by,
            remarks        = remarks,
        )

        # deduct from flock current count
        flock.current_count -= quantity
        flock.save()

        messages.success(request, f'Sale recorded! Revenue: ₱{total_revenue}')
    return redirect('growing_sales')

@login_required
def growing_sale_delete(request, pk):
    sale  = get_object_or_404(GrowingSale, pk=pk)
    # restore flock count
    flock = sale.flock
    if flock:
        flock.current_count += sale.quantity
        flock.save()
    sale.delete()
    messages.success(request, 'Sale record deleted and flock count restored!')
    return redirect('growing_sales')

@login_required
def laying_sales_list(request):
    sales     = LayingSale.objects.all().order_by('-sale_date')
    buildings = Building.objects.filter(type='Laying')
    customers = Customer.objects.all()
    price_obj = EggPriceConfig.objects.first()
    total     = float(sales.aggregate(Sum('total_revenue'))['total_revenue__sum'] or 0)
    return render(request, 'sales/laying.html', {
        'sales':         sales,
        'buildings':     buildings,
        'customers':     customers,
        'current_price': price_obj,
        'total':         total,
    })

@login_required
def laying_sale_save(request):
    if request.method == 'POST':
        building      = request.POST['building']
        customer_id   = request.POST.get('customer_id') or None
        eggs_sold     = int(request.POST['eggs_sold'])
        price_per_egg = float(request.POST['price_per_egg'])
        total_revenue = round(eggs_sold * price_per_egg, 2)
        sale_date     = request.POST['sale_date']
        recorded_by   = request.POST['recorded_by']
        remarks       = request.POST.get('remarks', '')

        LayingSale.objects.create(
            building      = building,
            customer_id   = customer_id,
            eggs_sold     = eggs_sold,
            price_per_egg = price_per_egg,
            total_revenue = total_revenue,
            sale_date     = sale_date,
            recorded_by   = recorded_by,
            remarks       = remarks,
        )
        messages.success(request, f'Sale recorded! Revenue: ₱{total_revenue}')
    return redirect('laying_sales')

@login_required
def laying_sale_delete(request, pk):
    sale = get_object_or_404(LayingSale, pk=pk)
    sale.delete()
    messages.success(request, 'Sale record deleted!')
    return redirect('laying_sales')

@login_required
def sales_analytics(request):
    from erp_config.models import SalesTarget
    from datetime import date

    today = date.today()

    # ── Growing Analytics (all houses combined) ───────────────
    # get active growing target (period covers today)
    growing_target_obj = SalesTarget.objects.filter(
        type='Growing',
        period_start__lte=today,
        period_end__gte=today
    ).first()

    growing_target       = float(growing_target_obj.target_revenue) if growing_target_obj else 0
    growing_period_start = growing_target_obj.period_start if growing_target_obj else None
    growing_period_end   = growing_target_obj.period_end   if growing_target_obj else None

    # filter sales and expenses within target period
    growing_sales_qs = GrowingSale.objects.all()
    growing_finance_qs = Finance.objects.all()
    if growing_target_obj:
        growing_sales_qs   = growing_sales_qs.filter(
            sale_date__gte=growing_period_start,
            sale_date__lte=growing_period_end)
        growing_finance_qs = growing_finance_qs.filter(
            expense_date__gte=growing_period_start,
            expense_date__lte=growing_period_end)

    growing_revenue  = float(growing_sales_qs.aggregate(Sum('total_revenue'))['total_revenue__sum'] or 0)
    growing_expenses = float(growing_finance_qs.aggregate(Sum('amount'))['amount__sum'] or 0)
    growing_net      = growing_revenue - growing_expenses
    growing_pct      = round((growing_revenue / growing_target * 100), 1) if growing_target > 0 else 0

    growing_data = [{
        'name':         'All Growing Houses',
        'revenue':      growing_revenue,
        'expenses':     growing_expenses,
        'net':          growing_net,
        'target':       growing_target,
        'pct':          min(growing_pct, 100),
        'gap':          max(growing_target - growing_revenue, 0),
        'period_start': growing_period_start,
        'period_end':   growing_period_end,
    }]

    # ── Laying Analytics (per building) ───────────────────────
    laying_buildings = Building.objects.filter(type='Laying')
    laying_data      = []

    for b in laying_buildings:
        # get active laying target for this building
        target_obj = SalesTarget.objects.filter(
            type='Laying',
            building=b.name,
            period_start__lte=today,
            period_end__gte=today
        ).first()

        target       = float(target_obj.target_revenue) if target_obj else 0
        period_start = target_obj.period_start if target_obj else None
        period_end   = target_obj.period_end   if target_obj else None

        sales_qs   = LayingSale.objects.filter(building=b.name)
        finance_qs = LayingFinance.objects.filter(building=b.name)

        if target_obj:
            sales_qs   = sales_qs.filter(
                sale_date__gte=period_start,
                sale_date__lte=period_end)
            finance_qs = finance_qs.filter(
                expense_date__gte=period_start,
                expense_date__lte=period_end)

        revenue  = float(sales_qs.aggregate(Sum('total_revenue'))['total_revenue__sum'] or 0)
        expenses = float(finance_qs.aggregate(Sum('amount'))['amount__sum'] or 0)
        net      = revenue - expenses
        pct      = round((revenue / target * 100), 1) if target > 0 else 0

        laying_data.append({
            'name':         b.name,
            'revenue':      revenue,
            'expenses':     expenses,
            'net':          net,
            'target':       target,
            'pct':          min(pct, 100),
            'gap':          max(target - revenue, 0),
            'period_start': period_start,
            'period_end':   period_end,
        })

    return render(request, 'sales/sales_analytics.html', {
        'growing_data': growing_data,
        'laying_data':  laying_data,
    })

@login_required
def export_growing_sales(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Growing Sales'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='17A98A')

    headers = ['Date', 'Batch', 'Growing House', 'Customer', 'Quantity', 'Price/Head', 'Total Revenue', 'Recorded By', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, s in enumerate(GrowingSale.objects.all().order_by('-sale_date'), 2):
        ws.cell(row=row, column=1, value=str(s.sale_date))
        ws.cell(row=row, column=2, value=s.flock.batch_name if s.flock else '—')
        ws.cell(row=row, column=3, value=s.flock.growing_house if s.flock else '—')
        ws.cell(row=row, column=4, value=s.customer.name if s.customer else '—')
        ws.cell(row=row, column=5, value=s.quantity)
        ws.cell(row=row, column=6, value=float(s.price_per_head))
        ws.cell(row=row, column=7, value=float(s.total_revenue))
        ws.cell(row=row, column=8, value=s.recorded_by)
        ws.cell(row=row, column=9, value=s.remarks)

    for col, width in enumerate([12, 18, 18, 18, 10, 12, 14, 14, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="growing_sales_{date.today()}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_laying_sales(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laying Sales'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='D4880A')

    headers = ['Date', 'Building', 'Customer', 'Eggs Sold', 'Price/Egg', 'Total Revenue', 'Recorded By', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, s in enumerate(LayingSale.objects.all().order_by('-sale_date'), 2):
        ws.cell(row=row, column=1, value=str(s.sale_date))
        ws.cell(row=row, column=2, value=s.building)
        ws.cell(row=row, column=3, value=s.customer.name if s.customer else '—')
        ws.cell(row=row, column=4, value=s.eggs_sold)
        ws.cell(row=row, column=5, value=float(s.price_per_egg))
        ws.cell(row=row, column=6, value=float(s.total_revenue))
        ws.cell(row=row, column=7, value=s.recorded_by)
        ws.cell(row=row, column=8, value=s.remarks)

    for col, width in enumerate([12, 18, 18, 12, 12, 14, 14, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="laying_sales_{date.today()}.xlsx"'
    wb.save(response)
    return response

@login_required
def test_dr_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER
    from django.http import HttpResponse
    import io
    import os

    buffer   = io.BytesIO()
    doc      = SimpleDocTemplate(buffer, pagesize=A4,
                                 rightMargin=15*mm, leftMargin=15*mm,
                                 topMargin=15*mm, bottomMargin=15*mm)
    elements = []

    # ── Colors ────────────────────────────────────────────────
    navy       = colors.HexColor('#1A3A6B')
    white      = colors.white
    lightgray  = colors.HexColor('#F5F5F5')
    light_blue = colors.HexColor('#E8EEF7')

    # ── Styles ────────────────────────────────────────────────
    farm_name_style = ParagraphStyle('farm_name', fontSize=11, fontName='Helvetica-Bold', textColor=colors.black, spaceAfter=3)
    farm_info_style = ParagraphStyle('farm_info', fontSize=8,  fontName='Helvetica',      textColor=colors.gray,  leading=13)
    dr_title_style  = ParagraphStyle('dr_title',  fontSize=20, fontName='Helvetica-Bold', textColor=navy, alignment=TA_RIGHT, charSpace=2)
    dr_sub_style    = ParagraphStyle('dr_sub',    fontSize=20, fontName='Helvetica-Bold', textColor=navy, alignment=TA_RIGHT, charSpace=2)
    label_style     = ParagraphStyle('label',     fontSize=8,  fontName='Helvetica-Bold', textColor=navy)
    small_style     = ParagraphStyle('small',     fontSize=8,  fontName='Helvetica',      textColor=colors.gray)
    small_bold      = ParagraphStyle('small_bold',fontSize=8,  fontName='Helvetica-Bold', textColor=colors.black)
    notes_style     = ParagraphStyle('notes',     fontSize=8,  fontName='Helvetica',      textColor=colors.gray,  leading=13)
    bold_style      = ParagraphStyle('bold',      fontSize=9,  fontName='Helvetica-Bold', textColor=colors.black)
    customer_style  = ParagraphStyle('customer',  fontSize=13, fontName='Helvetica-Bold', textColor=colors.black, spaceAfter=2)

    # ── Logo ──────────────────────────────────────────────────
    logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'logo.jpg')

    # ── Header ────────────────────────────────────────────────
    farm_info_cell = [
        Paragraph("3C's Farm", farm_name_style),
        Paragraph("Balagtasin 2nd, San Jose Batangas", farm_info_style),
        Paragraph("3csfarm@gmail.com", farm_info_style),
        Paragraph("09620509923", farm_info_style),
    ]

    if os.path.exists(logo_path):
        logo = Image(logo_path, width=25*mm, height=25*mm)
        logo.hAlign = 'CENTER'
        logo_cell = [logo]
    else:
        logo_cell = [Paragraph("3C's Farm", farm_name_style)]

    logo_table = Table([[logo_cell]], colWidths=[33*mm])
    logo_table.setStyle(TableStyle([
        ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))

    dr_title_cell = [
        Spacer(1, 4*mm),
        Paragraph("DELIVERY", dr_title_style),
        Spacer(1, 1*mm),
        Paragraph("RECEIPT",  dr_sub_style),
    ]

    header_table = Table(
        [[farm_info_cell, logo_table, dr_title_cell]],
        colWidths=[65*mm, 35*mm, 80*mm]
    )
    header_table.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('ALIGN',        (1,0), (1,0),   'CENTER'),
        ('ALIGN',        (2,0), (2,0),   'RIGHT'),
        ('LEFTPADDING',  (1,0), (1,0),   6),
        ('RIGHTPADDING', (1,0), (1,0),   6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 6*mm))
    elements.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#CCCCCC')))
    elements.append(Spacer(1, 6*mm))

    # ── Billed To + DR Info ───────────────────────────────────
    billed_cell = [
        Paragraph("Billed To", label_style),
        Spacer(1, 3*mm),
        Paragraph("Juan dela Cruz", customer_style),
        Paragraph("Wholesaler", small_style),
    ]

    dr_info_data = [
        [Paragraph("Receipt #",    ParagraphStyle('rl', fontSize=8, fontName='Helvetica-Bold', textColor=navy)),
         Paragraph("DR-2026-001",  small_bold)],
        [Paragraph("Receipt date", ParagraphStyle('rl', fontSize=8, fontName='Helvetica-Bold', textColor=navy)),
         Paragraph("May 24, 2026", small_bold)],
        [Paragraph("Copy",         ParagraphStyle('rl', fontSize=8, fontName='Helvetica-Bold', textColor=navy)),
         Paragraph("ORIGINAL COPY", ParagraphStyle('orig', fontSize=8, fontName='Helvetica-Bold', textColor=colors.HexColor('#2ECC71')))],
    ]
    dr_info_table = Table(dr_info_data, colWidths=[30*mm, 50*mm])
    dr_info_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0),  light_blue),
        ('BACKGROUND',    (0,2), (-1,2),  light_blue),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
        ('LINEBELOW',     (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
    ]))

    billing_table = Table(
        [[billed_cell, dr_info_table]],
        colWidths=[95*mm, 85*mm]
    )
    billing_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN',  (1,0), (1,0),   'RIGHT'),
    ]))
    elements.append(billing_table)
    elements.append(Spacer(1, 8*mm))

    # ── Items Table ───────────────────────────────────────────
    item_data = [['QTY', 'Description', 'Unit Price', 'Amount']]
    items     = [(500, 'Laying Chicken', 150.00)]

    for qty, desc, price in items:
        item_data.append([str(qty), desc, f'P{price:,.2f}', f'P{qty*price:,.2f}'])

    items_table = Table(item_data, colWidths=[20*mm, 90*mm, 35*mm, 35*mm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0),  navy),
        ('TEXTCOLOR',     (0,0), (-1,0),  white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 9),
        ('ALIGN',         (0,0), (-1,0),  'LEFT'),
        ('ALIGN',         (0,1), (0,-1),  'CENTER'),
        ('ALIGN',         (2,1), (3,-1),  'RIGHT'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [white, lightgray]),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('TOPPADDING',    (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 2*mm))

    # ── Totals ────────────────────────────────────────────────
    subtotal    = sum(qty * price for qty, _, price in items)
    totals_data = [
        ['', '', 'Subtotal',    f'P{subtotal:,.2f}'],
        ['', '', 'TOTAL (PHP)', f'P{subtotal:,.2f}'],
    ]
    totals_table = Table(totals_data, colWidths=[20*mm, 90*mm, 35*mm, 35*mm])
    totals_table.setStyle(TableStyle([
        ('ALIGN',         (2,0), (3,-1),  'RIGHT'),
        ('FONTNAME',      (2,0), (3,0),   'Helvetica'),
        ('FONTNAME',      (2,1), (3,1),   'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 9),
        ('TEXTCOLOR',     (2,1), (3,1),   white),
        ('BACKGROUND',    (2,1), (3,1),   navy),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
        ('LINEABOVE',     (2,0), (3,0),   0.5, colors.HexColor('#DDDDDD')),
    ]))
    elements.append(totals_table)
    elements.append(Spacer(1, 14*mm))

    # ── Signature Lines ───────────────────────────────────────
    sig_data = [
        [Paragraph('Prepared by:', small_style),
         Paragraph('Received by:', small_style),
         Paragraph('Noted by:', small_style)],
        [Paragraph('___________________', small_style),
         Paragraph('___________________', small_style),
         Paragraph('___________________', small_style)],
        [Paragraph('Staff / Encoder', small_style),
         Paragraph('Customer Signature', small_style),
         Paragraph('Manager / Owner', small_style)],
    ]
    sig_table = Table(sig_data, colWidths=[60*mm, 60*mm, 60*mm])
    sig_table.setStyle(TableStyle([
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1, 10*mm))

    # ── Notes ─────────────────────────────────────────────────
    elements.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#CCCCCC')))
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph("Notes", label_style))
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(
        "Thank you for your purchase! All sales are final. "
        "Please retain this receipt as proof of transaction. "
        "For questions or support, contact us at 3csfarm@gmail.com or 09620509923.",
        notes_style
    ))

    # ── Build ─────────────────────────────────────────────────
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="DR-2026-001.pdf"'
    return response

@login_required
def sales_analytics_data(request):
    from django.http import JsonResponse
    from erp_config.models import SalesTarget, Building
    from datetime import date

    today = date.today()

    # ── Growing ───────────────────────────────────────────────
    growing_target_obj = SalesTarget.objects.filter(
        type='Growing',
        period_start__lte=today,
        period_end__gte=today
    ).first()

    growing_target  = float(growing_target_obj.target_revenue) if growing_target_obj else 0
    growing_sales_qs = GrowingSale.objects.all()
    if growing_target_obj:
        growing_sales_qs = growing_sales_qs.filter(
            sale_date__gte=growing_target_obj.period_start,
            sale_date__lte=growing_target_obj.period_end
        )
    growing_revenue = float(growing_sales_qs.aggregate(
                        Sum('total_revenue'))['total_revenue__sum'] or 0)
    growing_pct     = round((growing_revenue / growing_target * 100), 1) if growing_target > 0 else 0

    growing = [{
        'name':         'All Growing Houses',
        'pct':          min(growing_pct, 100),
        'period_start': str(growing_target_obj.period_start) if growing_target_obj else None,
        'period_end':   str(growing_target_obj.period_end)   if growing_target_obj else None,
    }]

    # ── Laying ────────────────────────────────────────────────
    laying_buildings = Building.objects.filter(type='Laying')
    laying = []
    for b in laying_buildings:
        target_obj = SalesTarget.objects.filter(
            type='Laying',
            building=b.name,
            period_start__lte=today,
            period_end__gte=today
        ).first()

        target = float(target_obj.target_revenue) if target_obj else 0
        sales_qs = LayingSale.objects.filter(building=b.name)
        if target_obj:
            sales_qs = sales_qs.filter(
                sale_date__gte=target_obj.period_start,
                sale_date__lte=target_obj.period_end
            )
        revenue = float(sales_qs.aggregate(
                    Sum('total_revenue'))['total_revenue__sum'] or 0)
        pct     = round((revenue / target * 100), 1) if target > 0 else 0

        laying.append({
            'name':         b.name,
            'pct':          min(pct, 100),
            'period_start': str(target_obj.period_start) if target_obj else None,
            'period_end':   str(target_obj.period_end)   if target_obj else None,
        })

    return JsonResponse({'growing': growing, 'laying': laying})