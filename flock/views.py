from urllib import request

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum
from .models import Flock, Mortality, FlockSnapshot
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.http import JsonResponse
import json



def auto_snapshot(flock):
    today      = date.today()
    days_old   = (today - flock.date_placed).days
    weeks_due  = days_old // 7

    if weeks_due == 0:
        return

    existing = FlockSnapshot.objects.filter(flock=flock).count()

    if existing >= weeks_due:
        return

    for week in range(existing + 1, weeks_due + 1):
        week_start = flock.date_placed + timedelta(days=(week-1)*7)
        week_end   = flock.date_placed + timedelta(days=week*7 - 1)

        deaths_this_week = Mortality.objects.filter(
            flock=flock,
            death_date__range=[week_start, week_end]
        ).aggregate(Sum('count'))['count__sum'] or 0

        total_deaths = Mortality.objects.filter(
            flock=flock,
            death_date__lte=week_end
        ).aggregate(Sum('count'))['count__sum'] or 0

        count_at_snapshot = flock.start_count - total_deaths

        FlockSnapshot.objects.create(
            flock=flock,
            batch_name=flock.batch_name,
            growing_house=flock.growing_house,
            week_number=week,
            snapshot_date=week_end,
            start_count=flock.start_count,
            deaths_this_week=deaths_this_week,
            count_at_snapshot=count_at_snapshot
        )

@login_required
def flock_list(request):
    from erp_config.models import Building, Cause
    from master_data.models import Supplier
    flocks_all  = Flock.objects.exclude(status='Sold').order_by('-Date')
    mort_all    = Mortality.objects.filter(flock__status__in=['Active', 'Transferred']).order_by('-death_date')
    sold_flocks    = Flock.objects.filter(status='Sold').order_by('-Date')
    sold_mortalities = Mortality.objects.filter(flock__status='Sold').order_by('-death_date')
    pag_flock   = Paginator(flocks_all, 10)
    pag_mort    = Paginator(mort_all, 10)
    flock_page  = request.GET.get('flock_page')
    mort_page   = request.GET.get('mort_page')
    flocks      = pag_flock.get_page(flock_page)
    mortalities = pag_mort.get_page(mort_page)
    for flock in flocks_all:
        if flock.date_placed:
            flock.age_days = (date.today() - flock.date_placed).days
            auto_snapshot(flock)
        else:
            flock.age_days = 0
    for flock in flocks:
        if flock.date_placed:
            flock.age_days = (date.today() - flock.date_placed).days
        else:
            flock.age_days = 0
    snapshots = FlockSnapshot.objects.all().order_by('flock', 'week_number')
    buildings  = Building.objects.filter(type='Growing')
    causes     = Cause.objects.all()
    suppliers  = Supplier.objects.all()
    return render(request, 'flock/list.html', {
        'flocks':           flocks,
        'mortalities':      mortalities,
        'snapshots':        snapshots,
        'buildings':        buildings,
        'causes':           causes,
        'suppliers':        suppliers,
        'sold_flocks':      sold_flocks,
        'sold_mortalities': sold_mortalities,
    })

@login_required
def flock_save(request):
    if request.method == 'POST':
        batch     = request.POST['batch_name']
        house     = request.POST['growing_house']
        start     = int(request.POST['start_count'])
        placed    = request.POST['date_placed']
        supplier  = request.POST.get('supplier', '')
        notes     = request.POST.get('notes', '')

        Flock.objects.create(
            batch_name=batch, growing_house=house,
            start_count=start, current_count=start,
            date_placed=placed, supplier=supplier,
            status='Active', notes=notes
        )
        messages.success(request, f'Flock "{batch}" added successfully!')
    return redirect('flock_list')

@login_required
def flock_delete(request, pk):
    flock = get_object_or_404(Flock, pk=pk)
    Mortality.objects.filter(flock=flock).delete()
    flock.delete()
    messages.success(request, 'Flock and its mortality records deleted!')
    return redirect('flock_list')

@login_required
def flock_transfer(request, pk):
    flock = get_object_or_404(Flock, pk=pk)
    if flock.growing_house == 'Growing House 2':
        flock.growing_house = 'Growing House 1'
        flock.status        = 'Transferred'
        flock.save()
        messages.success(request, f'"{flock.batch_name}" transferred to Growing House 1!')
    else:
        messages.error(request, 'This flock is already in Growing House 1')
    return redirect('flock_list')

@login_required
def mortality_save(request):
    if request.method == 'POST':
        flock_id   = request.POST['flock_id']
        house      = request.POST['growing_house']
        death_date = request.POST['death_date']
        count      = int(request.POST['count'])
        cause      = request.POST.get('cause', '')
        recorded   = request.POST['recorded_by']
        remarks    = request.POST.get('remarks', '')

        flock = get_object_or_404(Flock, pk=flock_id)

        if count > flock.current_count:
            messages.error(request, f'Death count exceeds current flock count ({flock.current_count})')
            return redirect('flock_list')

        Mortality.objects.create(
            flock=flock, growing_house=house,
            death_date=death_date, count=count,
            cause=cause, recorded_by=recorded, remarks=remarks
        )

        flock.current_count -= count
        flock.save()

        messages.success(request, f'{count} deaths logged. Remaining: {flock.current_count}')
    return redirect('flock_list')

@login_required
def mortality_delete(request, pk):
    mortality = get_object_or_404(Mortality, pk=pk)
    flock     = mortality.flock
    flock.current_count += mortality.count
    flock.save()
    mortality.delete()
    messages.success(request, 'Mortality record deleted and flock count restored!')
    return redirect('flock_list')

@login_required
def flock_update(request, pk):
    from erp_config.models import Building
    from master_data.models import Supplier
    flock = get_object_or_404(Flock, pk=pk)
    if request.method == 'POST':
        flock.batch_name    = request.POST['batch_name']
        flock.growing_house = request.POST['growing_house']
        flock.start_count   = request.POST['start_count']
        flock.current_count = request.POST['current_count']
        flock.date_placed   = request.POST['date_placed']
        flock.supplier      = request.POST.get('supplier', '')
        flock.status        = request.POST['status']
        flock.notes         = request.POST.get('notes', '')
        flock.save()
        # delete old snapshots so they regenerate correctly
        FlockSnapshot.objects.filter(flock=flock).delete()
        messages.success(request, f'"{flock.batch_name}" updated successfully!')
        return redirect('flock_list')
    buildings = Building.objects.filter(type='Growing')
    suppliers = Supplier.objects.all()
    return render(request, 'flock/edit.html', {
        'flock':     flock,
        'buildings': buildings,
        'suppliers': suppliers,
    })

@login_required
def mortality_update(request, pk):
    mortality = get_object_or_404(Mortality, pk=pk)
    if request.method == 'POST':
        mortality.growing_house = request.POST['growing_house']
        mortality.death_date    = request.POST['death_date']
        mortality.count         = request.POST['count']
        mortality.cause         = request.POST.get('cause', '')
        mortality.recorded_by   = request.POST['recorded_by']
        mortality.remarks       = request.POST.get('remarks', '')
        mortality.save()
        messages.success(request, 'Mortality record updated!')
        return redirect('flock_list')
    flocks = Flock.objects.all()
    return render(request, 'flock/mortality_edit.html', {
        'mortality': mortality,
        'flocks':    flocks
    })

@login_required
def export_flock(request):
    wb = openpyxl.Workbook()

    # ── Sheet 1 — Flocks ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Flocks'

    header_font  = Font(bold=True, color='FFFFFF')
    flock_fill   = PatternFill(fill_type='solid', fgColor='2ECC71')
    mort_fill    = PatternFill(fill_type='solid', fgColor='E74C3C')

    flock_headers = ['Batch Name', 'Growing House', 'Start Count',
                     'Current Count', 'Date Placed', 'Supplier', 'Status', 'Age (days)']
    for col, header in enumerate(flock_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = flock_fill
        cell.alignment = Alignment(horizontal='center')

    for row, flock in enumerate(Flock.objects.all().order_by('-Date'), 2):
        age = (date.today() - flock.date_placed).days if flock.date_placed else 0
        ws1.cell(row=row, column=1, value=flock.batch_name)
        ws1.cell(row=row, column=2, value=flock.growing_house)
        ws1.cell(row=row, column=3, value=flock.start_count)
        ws1.cell(row=row, column=4, value=flock.current_count)
        ws1.cell(row=row, column=5, value=str(flock.date_placed))
        ws1.cell(row=row, column=6, value=flock.supplier)
        ws1.cell(row=row, column=7, value=flock.status)
        ws1.cell(row=row, column=8, value=age)

    for col, width in enumerate([18, 18, 12, 14, 14, 16, 12, 12], 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ── Sheet 2 — Mortality ───────────────────────────────────
    ws2 = wb.create_sheet(title='Mortality')

    mort_headers = ['Date', 'Batch', 'Growing House',
                    'Deaths', 'Cause', 'Recorded By', 'Remarks']
    for col, header in enumerate(mort_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = mort_fill
        cell.alignment = Alignment(horizontal='center')

    for row, m in enumerate(Mortality.objects.all().order_by('-death_date'), 2):
        ws2.cell(row=row, column=1, value=str(m.death_date))
        ws2.cell(row=row, column=2, value=m.flock.batch_name)
        ws2.cell(row=row, column=3, value=m.growing_house)
        ws2.cell(row=row, column=4, value=m.count)
        ws2.cell(row=row, column=5, value=m.cause)
        ws2.cell(row=row, column=6, value=m.recorded_by)
        ws2.cell(row=row, column=7, value=m.remarks)

    for col, width in enumerate([14, 18, 18, 10, 14, 16, 20], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ── Sheet 3 — Weekly Snapshots ────────────────────────────
    ws3 = wb.create_sheet(title='Weekly Snapshots')
    snap_fill = PatternFill(fill_type='solid', fgColor='1A8FD1')

    snap_headers = ['Week', 'Batch', 'Growing House', 'Snapshot Date',
                    'Start Count', 'Deaths This Week', 'Count at Snapshot']
    for col, header in enumerate(snap_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = snap_fill
        cell.alignment = Alignment(horizontal='center')

    for row, s in enumerate(FlockSnapshot.objects.all().order_by('flock', 'week_number'), 2):
        ws3.cell(row=row, column=1, value=s.week_number)
        ws3.cell(row=row, column=2, value=s.batch_name)
        ws3.cell(row=row, column=3, value=s.growing_house)
        ws3.cell(row=row, column=4, value=str(s.snapshot_date))
        ws3.cell(row=row, column=5, value=s.start_count)
        ws3.cell(row=row, column=6, value=s.deaths_this_week)
        ws3.cell(row=row, column=7, value=s.count_at_snapshot)

    for col, width in enumerate([8, 18, 18, 14, 12, 16, 16], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="flock_{date.today()}.xlsx"'
    wb.save(response)
    return response

def flock_info(request, pk):
    try:
        flock = Flock.objects.get(pk=pk)
        if flock.status == 'Sold':
            return JsonResponse({
                'found':  True,
                'sold':   True,
                'batch_name': flock.batch_name,
            })
        return JsonResponse({
            'found':         True,
            'sold':          False,
            'flock_id':      flock.pk,
            'batch_name':    flock.batch_name,
            'growing_house': flock.growing_house,
            'current_count': flock.current_count,
        })
    except Flock.DoesNotExist:
        return JsonResponse({'found': False})

@login_required
def quick_mortality(request):
    if request.method == 'POST':
        try:
            data         = json.loads(request.body)
            flock_id     = data.get('flock_id')
            count        = int(data.get('count', 0))
            cause        = data.get('cause', '')
            recorded_by  = data.get('recorded_by', '')
            death_date   = data.get('death_date')
            remarks      = data.get('remarks', '')
            growing_house = data.get('growing_house', '')

            flock = Flock.objects.get(pk=flock_id)

            if count > flock.current_count:
                return JsonResponse({
                    'success': False,
                    'error': f'Death count exceeds current count ({flock.current_count})'
                })

            Mortality.objects.create(
                flock         = flock,
                growing_house = growing_house,
                death_date    = death_date,
                count         = count,
                cause         = cause,
                recorded_by   = recorded_by,
                remarks       = remarks
            )

            flock.current_count -= count
            flock.save()

            return JsonResponse({
                'success': True,
                'message': f'{count} deaths logged. Remaining: {flock.current_count}'
            })

        except Exception as e:
            return JsonResponse({'success': False,'error': str(e)})

    return JsonResponse({'success': False,'error': 'Invalid request'})