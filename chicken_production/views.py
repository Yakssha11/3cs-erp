from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from .models import ChickenProduction
from flock.models import Flock
from erp_config.models import ChickenPriceConfig, Building
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@login_required
def chicken_production_list(request):
    productions   = ChickenProduction.objects.all().order_by('-harvest_date')
    flocks        = Flock.objects.filter(status='Active')
    current_price = ChickenPriceConfig.objects.first()
    buildings     = Building.objects.filter(type='Growing')

    paginator   = Paginator(productions, 10)
    page        = request.GET.get('page')
    productions = paginator.get_page(page)

    return render(request, 'chicken_production/list.html', {
        'productions':   productions,
        'flocks':        flocks,
        'current_price': current_price,
        'buildings':     buildings,
    })

@login_required
def chicken_production_save(request):
    if request.method == 'POST':
        flock_id        = request.POST['flock_id']
        growing_house   = request.POST['growing_house']
        harvest_date    = request.POST['harvest_date']
        total_harvested = int(request.POST['total_harvested'])
        good_chickens   = int(request.POST['good_chickens'])
        recorded_by     = request.POST['recorded_by']
        remarks         = request.POST.get('remarks', '')

        price_obj = ChickenPriceConfig.objects.first()
        if not price_obj:
            messages.error(request, 'No chicken price set! Please set a price in Config first.')
            return redirect('chicken_production_list')

        flock = get_object_or_404(Flock, pk=flock_id)

        if total_harvested > flock.current_count:
            messages.error(request, f'Harvest count exceeds flock count ({flock.current_count})')
            return redirect('chicken_production_list')

        rejected      = total_harvested - good_chickens
        flock_count   = flock.current_count
        harvest_rate  = round((good_chickens / flock_count * 100), 2) if flock_count > 0 else 0
        price_chicken = price_obj.price_chicken
        total_revenue = round(good_chickens * float(price_chicken), 2)

        ChickenProduction.objects.create(
            flock_id        = flock_id,
            growing_house   = growing_house,
            harvest_date    = harvest_date,
            total_harvested = total_harvested,
            good_chickens   = good_chickens,
            rejected        = rejected,
            flock_count     = flock_count,
            harvest_rate    = harvest_rate,
            price_chicken   = price_chicken,
            total_revenue   = total_revenue,
            recorded_by     = recorded_by,
            remarks         = remarks
        )

        # deduct from flock
        flock.current_count -= total_harvested
        flock.save()

        messages.success(request, f'Harvest logged! Revenue: ₱{total_revenue}')
    return redirect('chicken_production_list')

@login_required
def chicken_production_delete(request, pk):
    production = get_object_or_404(ChickenProduction, pk=pk)
    # restore flock count
    flock = production.flock
    if flock:
        flock.current_count += production.total_harvested
        flock.save()
    production.delete()
    messages.success(request, 'Record deleted and flock count restored!')
    return redirect('chicken_production_list')

@login_required
def export_chicken_production(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Chicken Production'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='17A98A')

    headers = ['Date', 'Growing House', 'Total Harvested', 'Good',
               'Rejected', 'Flock Count', 'Harvest Rate %',
               'Price/Chicken', 'Revenue', 'Recorded By', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, p in enumerate(ChickenProduction.objects.all().order_by('-harvest_date'), 2):
        ws.cell(row=row, column=1,  value=str(p.harvest_date))
        ws.cell(row=row, column=2,  value=p.growing_house)
        ws.cell(row=row, column=3,  value=p.total_harvested)
        ws.cell(row=row, column=4,  value=p.good_chickens)
        ws.cell(row=row, column=5,  value=p.rejected)
        ws.cell(row=row, column=6,  value=p.flock_count)
        ws.cell(row=row, column=7,  value=float(p.harvest_rate))
        ws.cell(row=row, column=8,  value=float(p.price_chicken))
        ws.cell(row=row, column=9,  value=float(p.total_revenue))
        ws.cell(row=row, column=10, value=p.recorded_by)
        ws.cell(row=row, column=11, value=p.remarks)

    for col, width in enumerate([12, 18, 16, 10, 10, 12, 14, 14, 14, 14, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="chicken_production_{date.today()}.xlsx"'
    wb.save(response)
    return response