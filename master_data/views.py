from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Customer, Supplier, Material
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
import random
import string

def generate_material_id():
    nums   = ''.join([str(random.randint(0,9)) for _ in range(3)])
    letter = random.choice(string.ascii_uppercase)
    return f"{nums}-{letter}"

@login_required
def master_data_view(request):
    from erp_config.models import Category, Unit
    customers  = Customer.objects.all()
    suppliers  = Supplier.objects.all()
    materials  = Material.objects.all()
    categories = Category.objects.all()
    units      = Unit.objects.all()
    return render(request, 'master_data/master_data.html', {
        'customers':  customers,
        'suppliers':  suppliers,
        'materials':  materials,
        'categories': categories,
        'units':      units,
        'active_tab': request.GET.get('tab', 'customers'),
    })

# ── Customers ─────────────────────────────────────────────
@login_required
def customer_save(request):
    if request.method == 'POST':
        name    = request.POST.get('name', '').strip()
        contact = request.POST.get('contact', '').strip()
        address = request.POST.get('address', '').strip()
        type    = request.POST.get('type', '')
        if name:
            Customer.objects.create(
                name    = name,
                contact = contact,
                address = address,
                type    = type,
            )
            messages.success(request, f'Customer "{name}" added!')
        else:
            messages.error(request, 'Customer name is required.')
    return redirect('/masterdata/?tab=customers')

@login_required
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    customer.delete()
    messages.success(request, f'Customer "{customer.name}" deleted!')
    return redirect('/masterdata/?tab=customers')

# ── Suppliers ─────────────────────────────────────────────
@login_required
def supplier_save(request):
    if request.method == 'POST':
        name    = request.POST.get('name', '').strip()
        contact = request.POST.get('contact', '').strip()
        address = request.POST.get('address', '').strip()
        type    = request.POST.get('type', '')
        if name:
            Supplier.objects.create(
                name    = name,
                contact = contact,
                address = address,
                type    = type,
            )
            messages.success(request, f'Supplier "{name}" added!')
        else:
            messages.error(request, 'Supplier name is required.')
    return redirect('/masterdata/?tab=suppliers')

@login_required
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    supplier.delete()
    messages.success(request, f'Supplier "{supplier.name}" deleted!')
    return redirect('/masterdata/?tab=suppliers')

# ── Materials ─────────────────────────────────────────────
@login_required
def material_save(request):
    if request.method == 'POST':
        from erp_config.models import Category, Unit
        name        = request.POST.get('name', '').strip()
        category    = request.POST.get('category', '')
        unit        = request.POST.get('unit', '')
        price       = request.POST.get('price', '0')
        description = request.POST.get('description', '')

        if name:
            # generate unique item_id
            item_id = generate_material_id()
            while Material.objects.filter(item_id=item_id).exists():
                item_id = generate_material_id()

            Material.objects.create(
                item_id     = item_id,
                name        = name,
                category    = category,
                unit        = unit,
                price       = price,
                description = description,
            )
            messages.success(request, f'Material "{name}" added with ID {item_id}!')
        else:
            messages.error(request, 'Material name is required.')
    return redirect('/masterdata/?tab=materials')

@login_required
def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)
    material.delete()
    messages.success(request, f'Material "{material.name}" deleted!')
    return redirect('/masterdata/?tab=materials')

@login_required
def get_materials(request):
    materials = Material.objects.all().values('item_id', 'name', 'category', 'unit', 'price')
    return JsonResponse({'materials': list(materials)})

@login_required
def master_data_records(request):
    active_tab = request.GET.get('tab', 'customers')
    search     = request.GET.get('search', '')

    # customers
    customers = Customer.objects.all().order_by('name')
    if search and active_tab == 'customers':
        customers = customers.filter(
            Q(name__icontains=search) |
            Q(contact__icontains=search) |
            Q(address__icontains=search) |
            Q(type__icontains=search)
        )
    pag_customers  = Paginator(customers, 10)
    customers_page = pag_customers.get_page(request.GET.get('cpage'))

    # suppliers
    suppliers = Supplier.objects.all().order_by('name')
    if search and active_tab == 'suppliers':
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(contact__icontains=search) |
            Q(address__icontains=search) |
            Q(type__icontains=search)
        )
    pag_suppliers  = Paginator(suppliers, 10)
    suppliers_page = pag_suppliers.get_page(request.GET.get('spage'))

    # export
    if 'export' in request.GET:
        wb = openpyxl.Workbook()

        header_font    = Font(bold=True, color='FFFFFF')
        header_fill    = PatternFill(start_color='0B1E2D', end_color='0B1E2D', fill_type='solid')
        header_align   = Alignment(horizontal='center')

        # customers sheet
        ws1 = wb.active
        ws1.title = 'Customers'
        ws1.append(['Name', 'Type', 'Contact', 'Address', 'Date Added'])
        for cell in ws1[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        for c in Customer.objects.all().order_by('name'):
            ws1.append([
                c.name,
                c.type,
                c.contact,
                c.address,
                c.Date.strftime('%Y-%m-%d'),
            ])

        # suppliers sheet
        ws2 = wb.create_sheet('Suppliers')
        ws2.append(['Name', 'Type', 'Contact', 'Address', 'Date Added'])
        for cell in ws2[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        for s in Supplier.objects.all().order_by('name'):
            ws2.append([
                s.name,
                s.type,
                s.contact,
                s.address,
                s.Date.strftime('%Y-%m-%d'),
            ])

        # materials sheet
        ws3 = wb.create_sheet('Materials')
        ws3.append(['Item ID', 'Name', 'Category', 'Unit', 'Price', 'Description', 'Date Added'])
        for cell in ws3[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        for m in Material.objects.all().order_by('name'):
            ws3.append([
                m.item_id,
                m.name,
                m.category,
                m.unit,
                float(m.price),
                m.description,
                m.Date.strftime('%Y-%m-%d'),
            ])

        # auto width
        for ws in [ws1, ws2, ws3]:
            for col in ws.columns:
                max_length = max((len(str(cell.value or '')) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="master_data.xlsx"'
        wb.save(response)
        return response

    # materials
    materials = Material.objects.all().order_by('name')
    if search and active_tab == 'materials':
        materials = materials.filter(
            Q(name__icontains=search) |
            Q(item_id__icontains=search) |
            Q(category__icontains=search)
        )
    pag_materials  = Paginator(materials, 10)
    materials_page = pag_materials.get_page(request.GET.get('mpage'))

    return render(request, 'master_data/records.html', {
        'customers':     customers_page,
        'suppliers':     suppliers_page,
        'materials':     materials_page,
        'active_tab':    active_tab,
        'search':        search,
    })