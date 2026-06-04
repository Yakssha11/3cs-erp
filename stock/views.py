from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum
from .models import Stock
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse, JsonResponse
from django.db.models import F

OWNER_USERNAME = 'CORONACIONR'

@login_required
def stock_list(request):
    from erp_config.models import Building
    from master_data.models import Material
    from django.db.models import Sum

    buildings = Building.objects.filter(type='Growing')
    materials = Material.objects.all()

    growing_building_names = Building.objects.filter(type='Growing').values_list('name', flat=True)
    stocks_raw = Stock.objects.filter(
        growing_house__in=growing_building_names
    ).order_by('item_id', F('expiry_date').asc(nulls_last=True), 'date')
    grouped    = {}
    today      = date.today()
    soon       = today + timedelta(days=30)

    for stock in stocks_raw:
        if stock.item_id not in grouped:
            grouped[stock.item_id] = {
                'item_id':      stock.item_id,
                'name':         stock.name,
                'category':     stock.category,
                'unit':         stock.unit,
                'price':        stock.price,
                'growing_house': stock.growing_house,
                'total_qty':    0,
                'batches':      [],
                'has_expiring': False,
                'has_expired':  False,
            }
        grouped[stock.item_id]['total_qty'] += int(stock.quantity)
        if stock.expiry_date:
            if stock.expiry_date < today:
                grouped[stock.item_id]['has_expired'] = True
            elif stock.expiry_date <= soon:
                grouped[stock.item_id]['has_expiring'] = True
        grouped[stock.item_id]['batches'].append(stock)

    grouped_list = list(grouped.values())
    is_owner     = request.user.username == OWNER_USERNAME

    return render(request, 'stock/list.html', {
        'grouped_stocks': grouped_list,
        'buildings':      buildings,
        'materials':      materials,
        'today':          today,
        'soon':           soon,
        'is_owner':       is_owner,
    })

@login_required
def stock_save(request):
    if request.method == 'POST':
        from master_data.models import Material
        item_id       = request.POST['item_id']
        quantity      = int(request.POST['quantity'])
        unit_quantity = request.POST.get('unit_quantity', '')
        growing_house = request.POST.get('growing_house', '')
        batch         = request.POST.get('batch', '')
        expiry_date   = request.POST.get('expiry_date') or None

        try:
            material = Material.objects.get(item_id=item_id)
        except Material.DoesNotExist:
            messages.error(request, 'Material not found!')
            if request.GET.get('next') == 'laying':
                return redirect('laying_stock')
            return redirect('stock_list')

        Stock.objects.create(
            item_id       = item_id,
            name          = material.name,
            price         = material.price,
            quantity      = quantity,
            unit_quantity = unit_quantity,
            category      = material.category,
            growing_house = growing_house,
            unit          = material.unit,
            batch         = batch,
            expiry_date   = expiry_date,
        )
        messages.success(request, f'Stock "{material.name}" — Batch {batch} added!')
    if request.GET.get('next') == 'laying':
        return redirect('laying_stock')
    return redirect('stock_list')

@login_required
def stock_delete(request, pk):
    from master_data.models import Material
    stock    = get_object_or_404(Stock, pk=pk)
    name     = stock.name
    batch    = stock.batch
    item_id  = stock.item_id
    stock.delete()

    # recalculate MAP after deletion
    try:
        material = Material.objects.get(item_id=item_id)
        material.recalculate_map()
    except Material.DoesNotExist:
        pass

    messages.success(request, f'"{name}" Batch {batch} deleted!')
    if request.GET.get('next') == 'laying':
        return redirect('laying_stock')
    return redirect('stock_list')

@login_required
def export_stock(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='196E78')

    headers = ['Item ID', 'Name', 'Batch', 'Unit Qty', 'Price',
               'Quantity', 'Category', 'Growing House', 'Unit',
               'Expiry Date', 'Date Added']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    from erp_config.models import Building
    if request.GET.get('next') == 'laying':
        building_names = Building.objects.filter(type='Laying').values_list('name', flat=True)
    else:
        building_names = Building.objects.filter(type='Growing').values_list('name', flat=True)

    for row, stock in enumerate(Stock.objects.filter(growing_house__in=building_names).order_by('item_id', 'expiry_date'), 2):
        ws.cell(row=row, column=1,  value=stock.item_id)
        ws.cell(row=row, column=2,  value=stock.name)
        ws.cell(row=row, column=3,  value=stock.batch)
        ws.cell(row=row, column=4,  value=stock.unit_quantity)
        ws.cell(row=row, column=5,  value=float(stock.price))
        ws.cell(row=row, column=6,  value=stock.quantity)
        ws.cell(row=row, column=7,  value=stock.category)
        ws.cell(row=row, column=8,  value=stock.growing_house)
        ws.cell(row=row, column=9,  value=stock.unit)
        ws.cell(row=row, column=10, value=str(stock.expiry_date) if stock.expiry_date else '')
        ws.cell(row=row, column=11, value=str(stock.date))

    for col, width in enumerate([12, 25, 12, 10, 12, 10, 18, 18, 10, 14, 22], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="stocks_{date.today()}.xlsx"'
    wb.save(response)
    return response

@login_required
def stock_update(request, pk):
    from erp_config.models import Building
    stock = get_object_or_404(Stock, pk=pk)
    if request.method == 'POST':
        stock.quantity      = request.POST['quantity']
        stock.unit_quantity = request.POST.get('unit_quantity', '')
        stock.batch         = request.POST.get('batch', '')
        stock.expiry_date   = request.POST.get('expiry_date') or None
        stock.growing_house = request.POST.get('growing_house', '')
        stock.save()
        messages.success(request, f'"{stock.name}" Batch {stock.batch} updated!')
        if request.POST.get('next') == 'laying':
            return redirect('laying_stock')
        return redirect('stock_list')
    buildings = Building.objects.filter(type='Growing') | Building.objects.filter(type='Laying')
    return render(request, 'stock/edit.html', {
        'stock':     stock,
        'buildings': buildings,
        'next':      request.GET.get('next', ''),
    })

@login_required
def laying_stock(request):
    from erp_config.models import Building
    from master_data.models import Material

    buildings = Building.objects.filter(type='Laying')
    materials = Material.objects.all()

    laying_building_names = buildings.values_list('name', flat=True)

    stocks_raw = Stock.objects.filter(
        growing_house__in=laying_building_names
    ).order_by('item_id', F('expiry_date').asc(nulls_last=True), 'date')

    grouped = {}
    today   = date.today()
    soon    = today + timedelta(days=30)

    for stock in stocks_raw:
        if stock.item_id not in grouped:
            grouped[stock.item_id] = {
                'item_id':       stock.item_id,
                'name':          stock.name,
                'category':      stock.category,
                'unit':          stock.unit,
                'price':         stock.price,
                'growing_house': stock.growing_house,
                'total_qty':     0,
                'batches':       [],
                'has_expiring':  False,
                'has_expired':   False,
            }
        grouped[stock.item_id]['total_qty'] += int(stock.quantity)
        if stock.expiry_date:
            if stock.expiry_date < today:
                grouped[stock.item_id]['has_expired'] = True
            elif stock.expiry_date <= soon:
                grouped[stock.item_id]['has_expiring'] = True
        grouped[stock.item_id]['batches'].append(stock)

    grouped_list = list(grouped.values())
    is_owner     = request.user.username == OWNER_USERNAME

    return render(request, 'stock/laying_list.html', {
        'grouped_stocks': grouped_list,
        'buildings':      buildings,
        'materials':      materials,
        'today':          today,
        'soon':           soon,
        'is_owner':       is_owner,
    })

@login_required
def get_items(request):
    building_name = request.GET.get('building')
    if not building_name:
        return JsonResponse({'items': []})

    batches = Stock.objects.filter(
        growing_house=building_name,
        quantity__gt=0
    ).order_by('item_id', F('expiry_date').asc(nulls_last=True))

    items = {}
    for stock in batches:
        if stock.item_id not in items:
            items[stock.item_id] = {
                'item_id':   stock.item_id,
                'name':      stock.name,
                'unit':      stock.unit,
                'category':  stock.category,
                'total_qty': 0,
                'batches':   [],
            }
        items[stock.item_id]['total_qty'] += int(stock.quantity)
        items[stock.item_id]['batches'].append({
            'id':     stock.pk,
            'batch':  stock.batch,
            'qty':    stock.quantity,
            'expiry': str(stock.expiry_date) if stock.expiry_date else None,
        })

    return JsonResponse({'items': list(items.values())})

@login_required
def stock_pricing(request):
    if request.user.username != OWNER_USERNAME:
        messages.error(request, 'Access denied.')
        return redirect('stock_list')

    from erp_config.models import Building
    tab = request.GET.get('tab', 'growing')

    growing_buildings = Building.objects.filter(type='Growing').values_list('name', flat=True)
    laying_buildings  = Building.objects.filter(type='Laying').values_list('name', flat=True)

    if tab == 'laying':
        stocks = Stock.objects.filter(
            growing_house__in=laying_buildings
        ).order_by('item_id', F('expiry_date').asc(nulls_last=True), 'date')
    else:
        stocks = Stock.objects.filter(
            growing_house__in=growing_buildings
        ).order_by('item_id', F('expiry_date').asc(nulls_last=True), 'date')

    unpriced_count = stocks.filter(unit_price__isnull=True).count()
    total_count    = stocks.count()

    growing_unpriced = Stock.objects.filter(
        growing_house__in=growing_buildings,
        unit_price__isnull=True
    ).count()
    laying_unpriced  = Stock.objects.filter(
        growing_house__in=laying_buildings,
        unit_price__isnull=True
    ).count()

    return render(request, 'stock/pricing.html', {
        'stocks':           stocks,
        'unpriced_count':   unpriced_count,
        'total_count':      total_count,
        'tab':              tab,
        'growing_unpriced': growing_unpriced,
        'laying_unpriced':  laying_unpriced,
    })

@login_required
def stock_price_save(request, pk):
    if request.user.username != OWNER_USERNAME:
        messages.error(request, 'Access denied.')
        return redirect('stock_list')

    if request.method == 'POST':
        from master_data.models import Material
        stock      = get_object_or_404(Stock, pk=pk)
        unit_price = request.POST.get('unit_price')

        if unit_price:
            stock.unit_price = float(unit_price)
            stock.save()

            # recalculate MAP for this material
            try:
                material = Material.objects.get(item_id=stock.item_id)
                material.recalculate_map()
            except Material.DoesNotExist:
                pass

            messages.success(request, f'Price set for {stock.name} Batch {stock.batch}!')

    return redirect('stock_pricing')