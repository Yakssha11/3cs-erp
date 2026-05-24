from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum
from .models import LayingFinance
from egg_production.models import EggProduction
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@login_required
def laying_finance_list(request):
    from erp_config.models import Building, Category
    period   = request.GET.get('period', 'all')
    finances = LayingFinance.objects.all().order_by('-expense_date')
    today    = date.today()

    if period == 'today':
        finances = finances.filter(expense_date=today)
    elif period == 'week':
        finances = finances.filter(expense_date__week=today.isocalendar()[1],
                                   expense_date__year=today.year)
    elif period == 'month':
        finances = finances.filter(expense_date__month=today.month,
                                   expense_date__year=today.year)
    elif period == 'year':
        finances = finances.filter(expense_date__year=today.year)
    elif period == 'custom':
        start = request.GET.get('start')
        end   = request.GET.get('end')
        if start and end:
            finances = finances.filter(expense_date__range=[start, end])

    total_expenses = float(sum(f.amount for f in finances))
    egg_revenue    = float(EggProduction.objects.aggregate(
                        total=Sum('total_revenue'))['total'] or 0)
    net_profit     = egg_revenue - total_expenses

    paginator  = Paginator(finances, 10)
    page       = request.GET.get('page')
    finances   = paginator.get_page(page)
    buildings  = Building.objects.filter(type='Laying')
    categories = Category.objects.all()

    return render(request, 'laying_finance/list.html', {
        'finances':       finances,
        'total_expenses': total_expenses,
        'egg_revenue':    egg_revenue,
        'net_profit':     net_profit,
        'period':         period,
        'buildings':      buildings,
        'categories':     categories,
    })

@login_required
def laying_finance_save(request):
    if request.method == 'POST':
        LayingFinance.objects.create(
            expense_date = request.POST['expense_date'],
            nature       = request.POST['nature'],
            building     = request.POST['building'],
            amount       = request.POST['amount'],
            remarks      = request.POST.get('remarks', ''),
            person       = request.POST['person'],
        )
        messages.success(request, 'Expense saved successfully!')
    return redirect('laying_finance_list')

@login_required
def laying_finance_delete(request, pk):
    finance = get_object_or_404(LayingFinance, pk=pk)
    finance.delete()
    messages.success(request, 'Record deleted!')
    return redirect('laying_finance_list')

@login_required
def laying_finance_update(request, pk):
    from erp_config.models import Building, Category
    finance = get_object_or_404(LayingFinance, pk=pk)
    if request.method == 'POST':
        finance.expense_date = request.POST['expense_date']
        finance.nature       = request.POST['nature']
        finance.building     = request.POST['building']
        finance.amount       = request.POST['amount']
        finance.remarks      = request.POST.get('remarks', '')
        finance.person       = request.POST['person']
        finance.save()
        messages.success(request, 'Record updated!')
        return redirect('laying_finance_list')
    buildings  = Building.objects.filter(type='Laying')
    categories = Category.objects.all()
    return render(request, 'laying_finance/edit.html', {
        'finance':    finance,
        'buildings':  buildings,
        'categories': categories,
    })

@login_required
def export_laying_finance(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laying Finance'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='D4880A')

    headers = ['Date', 'Nature', 'Building', 'Amount', 'Person', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, f in enumerate(LayingFinance.objects.all().order_by('-expense_date'), 2):
        ws.cell(row=row, column=1, value=str(f.expense_date))
        ws.cell(row=row, column=2, value=f.nature)
        ws.cell(row=row, column=3, value=f.building)
        ws.cell(row=row, column=4, value=float(f.amount))
        ws.cell(row=row, column=5, value=f.person)
        ws.cell(row=row, column=6, value=f.remarks)

    for col, width in enumerate([14, 16, 18, 12, 16, 25], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="laying_finance_{date.today()}.xlsx"'
    wb.save(response)
    return response