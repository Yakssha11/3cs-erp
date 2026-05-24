from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from .models import EggProduction
from laying_flock.models import LayingFlock
from erp_config.models import EggPriceConfig
from datetime import date

@login_required
def egg_production_list(request):
    from erp_config.models import Building
    productions   = EggProduction.objects.all().order_by('-collection_date')
    flocks        = LayingFlock.objects.filter(status='Active')
    current_price = EggPriceConfig.objects.first()
    buildings     = Building.objects.filter(type='Laying')

    paginator    = Paginator(productions, 10)
    page         = request.GET.get('page')
    productions  = paginator.get_page(page)

    return render(request, 'egg_production/list.html', {
        'productions':   productions,
        'flocks':        flocks,
        'current_price': current_price,
        'buildings':     buildings,
    })

@login_required
def egg_production_save(request):
    if request.method == 'POST':
        flock_id   = request.POST['flock_id']
        building   = request.POST['building']
        coll_date  = request.POST['collection_date']
        total_eggs = int(request.POST['total_eggs'])
        good_eggs  = int(request.POST['good_eggs'])
        recorded   = request.POST['recorded_by']
        remarks    = request.POST.get('remarks', '')

        price_obj = EggPriceConfig.objects.first()
        if not price_obj:
            messages.error(request, 'No egg price set! Please set a price in Config first.')
            return redirect('egg_production_list')

        flock = get_object_or_404(LayingFlock, pk=flock_id)

        cracked_eggs    = total_eggs - good_eggs
        hen_count       = flock.current_count
        production_rate = round((good_eggs / hen_count * 100), 2) if hen_count > 0 else 0
        price_per_egg   = price_obj.price_per_egg
        total_revenue   = round(good_eggs * float(price_per_egg), 2)

        EggProduction.objects.create(
            flock_id        = flock_id,
            building        = building,
            collection_date = coll_date,
            total_eggs      = total_eggs,
            good_eggs       = good_eggs,
            cracked_eggs    = cracked_eggs,
            hen_count       = hen_count,
            production_rate = production_rate,
            price_per_egg   = price_per_egg,
            total_revenue   = total_revenue,
            recorded_by     = recorded,
            remarks         = remarks
        )

        messages.success(request, f'Egg production logged! Revenue: ₱{total_revenue}')
    return redirect('egg_production_list')

@login_required
def egg_production_delete(request, pk):
    production = get_object_or_404(EggProduction, pk=pk)
    production.delete()
    messages.success(request, 'Record deleted!')
    return redirect('egg_production_list')

@login_required
def egg_production_update(request, pk):
    from erp_config.models import Building
    production = get_object_or_404(EggProduction, pk=pk)
    if request.method == 'POST':
        production.building        = request.POST['building']
        production.collection_date = request.POST['collection_date']
        production.total_eggs      = int(request.POST['total_eggs'])
        production.good_eggs       = int(request.POST['good_eggs'])
        production.cracked_eggs    = production.total_eggs - production.good_eggs
        production.recorded_by     = request.POST['recorded_by']
        production.remarks         = request.POST.get('remarks', '')
        production.production_rate = round(
            (production.good_eggs / production.hen_count * 100), 2
        ) if production.hen_count > 0 else 0
        production.total_revenue = round(
            production.good_eggs * float(production.price_per_egg), 2
        )
        production.save()
        messages.success(request, 'Record updated!')
        return redirect('egg_production_list')
    flocks    = LayingFlock.objects.filter(status='Active')
    buildings = Building.objects.filter(type='Laying')
    return render(request, 'egg_production/edit.html', {
        'production': production,
        'flocks':     flocks,
        'buildings':  buildings,
    })

@login_required
def export_eggs(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Egg Production'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='D4880A')

    headers = ['Date', 'Building', 'Total Eggs', 'Good Eggs',
               'Cracked', 'Hen Count', 'Production Rate %',
               'Price/Egg', 'Revenue', 'Recorded By', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, p in enumerate(EggProduction.objects.all().order_by('-collection_date'), 2):
        ws.cell(row=row, column=1,  value=str(p.collection_date))
        ws.cell(row=row, column=2,  value=p.building)
        ws.cell(row=row, column=3,  value=p.total_eggs)
        ws.cell(row=row, column=4,  value=p.good_eggs)
        ws.cell(row=row, column=5,  value=p.cracked_eggs)
        ws.cell(row=row, column=6,  value=p.hen_count)
        ws.cell(row=row, column=7,  value=float(p.production_rate))
        ws.cell(row=row, column=8,  value=float(p.price_per_egg))
        ws.cell(row=row, column=9,  value=float(p.total_revenue))
        ws.cell(row=row, column=10, value=p.recorded_by)
        ws.cell(row=row, column=11, value=p.remarks)

    for col, width in enumerate([12, 16, 10, 10, 10, 10, 16, 10, 12, 14, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="egg_production_{date.today()}.xlsx"'
    wb.save(response)
    return response