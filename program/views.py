from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Program, ProgramStep, ProgramBatch
from stock.models import Stock


def convert_to_program_unit(consumed_base, program_unit, material):
    base_unit         = material.base_unit or ''
    stock_unit        = material.stock_unit or ''
    mat_unit          = material.unit or ''
    conversion_factor = float(material.conversion_factor) if material.conversion_factor else None

    pu = (program_unit or '').lower().strip()
    bu = base_unit.lower().strip()
    su = stock_unit.lower().strip()
    mu = mat_unit.lower().strip()

    if pu == bu:
        return consumed_base, True

    if conversion_factor and (pu == su or pu == mu or pu.rstrip('s') == su.rstrip('s') or pu.rstrip('s') == mu.rstrip('s')):
        return consumed_base / conversion_factor, True

    METRIC = {
        ('g',  'kg'): 1000,
        ('kg', 'g' ): 0.001,
        ('ml', 'l' ): 1000,
        ('l',  'ml'): 0.001,
        ('mg', 'g' ): 1000,
        ('g',  'l' ): 1000,
        ('ml', 'kg'): 1000,
    }
    if (pu, bu) in METRIC:
        return consumed_base * METRIC[(pu, bu)], True

    return None, False


@login_required
def program_growing(request):
    from erp_config.models import Unit
    from master_data.models import Material
    from flock.models import Flock
    programs  = Program.objects.filter(type='Growing').prefetch_related('steps', 'batches')
    materials = Material.objects.all().order_by('name')
    units     = Unit.objects.all()
    flocks    = Flock.objects.filter(status='Active').order_by('batch_name')
    return render(request, 'program/growing.html', {
        'programs':  programs,
        'materials': materials,
        'units':     units,
        'flocks':    flocks,
    })


@login_required
def program_laying(request):
    from erp_config.models import Unit
    from master_data.models import Material
    from laying_flock.models import LayingFlock
    programs  = Program.objects.filter(type='Laying').prefetch_related('steps', 'batches')
    materials = Material.objects.all().order_by('name')
    units     = Unit.objects.all()
    flocks    = LayingFlock.objects.filter(status='Active').order_by('batch_name')
    return render(request, 'program/laying.html', {
        'programs':  programs,
        'materials': materials,
        'units':     units,
        'flocks':    flocks,
    })


@login_required
def program_save(request):
    if request.method == 'POST':
        Program.objects.create(
            name        = request.POST['name'],
            type        = request.POST['type'],
            description = request.POST.get('description', '')
        )
        messages.success(request, 'Program created!')
    return redirect(request.POST.get('next', 'program_growing'))


@login_required
def program_delete(request, pk):
    program = get_object_or_404(Program, pk=pk)
    ptype   = program.type
    program.delete()
    messages.success(request, 'Program deleted!')
    if ptype == 'Laying':
        return redirect('program_laying')
    return redirect('program_growing')


@login_required
def batch_save(request):
    if request.method == 'POST':
        program    = get_object_or_404(Program, pk=request.POST['program_id'])
        flock_id   = request.POST['flock_id']
        flock_type = request.POST['flock_type']
        date_started = request.POST['date_started']
        notes      = request.POST.get('notes', '')

        ProgramBatch.objects.create(
            program      = program,
            flock_id     = flock_id,
            flock_type   = flock_type,
            date_started = date_started,
            notes        = notes,
        )
        messages.success(request, 'Batch assigned to program!')
    if program.type == 'Laying':
        return redirect('program_laying')
    return redirect('program_growing')


@login_required
def batch_delete(request, pk):
    batch  = get_object_or_404(ProgramBatch, pk=pk)
    ptype  = batch.program.type
    batch.delete()
    messages.success(request, 'Batch removed from program!')
    if ptype == 'Laying':
        return redirect('program_laying')
    return redirect('program_growing')


@login_required
def step_save(request, program_pk):
    program = get_object_or_404(Program, pk=program_pk)
    if request.method == 'POST':
        ProgramStep.objects.create(
            program            = program,
            cycle              = request.POST.get('cycle', 1),
            week               = request.POST['week'],
            day                = request.POST['day'],
            medicine           = request.POST.get('medicine', ''),
            dose_amount        = request.POST.get('dose_amount') or None,
            dose_unit          = request.POST.get('dose_unit', ''),
            dose_per           = request.POST.get('dose_per', ''),
            method             = request.POST.get('method', ''),
            remarks            = request.POST.get('remarks', ''),
            feed               = request.POST.get('feed', ''),
            feed_rate_per_bird = request.POST.get('feed_rate_per_bird') or None,
            feed_unit          = request.POST.get('feed_unit', ''),
        )
        messages.success(request, 'Step added!')
    if program.type == 'Laying':
        return redirect('program_laying')
    return redirect('program_growing')


@login_required
def step_delete(request, pk):
    step  = get_object_or_404(ProgramStep, pk=pk)
    ptype = step.program.type
    step.delete()
    messages.success(request, 'Step deleted!')
    if ptype == 'Laying':
        return redirect('program_laying')
    return redirect('program_growing')


@login_required
def step_update(request, pk):
    step = get_object_or_404(ProgramStep, pk=pk)
    if request.method == 'POST':
        step.cycle              = request.POST.get('cycle', 1)
        step.week               = request.POST['week']
        step.day                = request.POST['day']
        step.medicine           = request.POST.get('medicine', '')
        step.dose_amount        = request.POST.get('dose_amount') or None
        step.dose_unit          = request.POST.get('dose_unit', '')
        step.dose_per           = request.POST.get('dose_per', '')
        step.method             = request.POST.get('method', '')
        step.remarks            = request.POST.get('remarks', '')
        step.feed               = request.POST.get('feed', '')
        step.feed_rate_per_bird = request.POST.get('feed_rate_per_bird') or None
        step.feed_unit          = request.POST.get('feed_unit', '')
        step.save()
        messages.success(request, 'Step updated!')
    if step.program.type == 'Laying':
        return redirect('program_laying')
    return redirect('program_growing')


@login_required
def export_program(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from master_data.models import Material
    from flock.models import Flock
    from laying_flock.models import LayingFlock
    from django.db.models import Sum

    program = get_object_or_404(Program, pk=pk)

    if program.type == 'Growing':
        total_birds = Flock.objects.filter(status='Active').aggregate(
            Sum('current_count'))['current_count__sum'] or 0
    else:
        total_birds = LayingFlock.objects.filter(status='Active').aggregate(
            Sum('current_count'))['current_count__sum'] or 0

    material_map = {m.item_id: m.name for m in Material.objects.all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = program.name[:31]

    header_font = Font(bold=True, color='FFFFFF', size=10)
    center      = Alignment(horizontal='center', vertical='center')
    left        = Alignment(horizontal='left',   vertical='center')

    blue_fill   = PatternFill(fill_type='solid', fgColor='2E75B6')
    purple_fill = PatternFill(fill_type='solid', fgColor='7030A0')
    orange_fill = PatternFill(fill_type='solid', fgColor='C55A11')
    green_fill  = PatternFill(fill_type='solid', fgColor='375623')

    headers = [
        ('Cycle #',                blue_fill),
        ('Age Display',            blue_fill),
        ('Population',             blue_fill),
        ('Medicine Name',          purple_fill),
        ('Medicine Item ID',       purple_fill),
        ('UoM',                    purple_fill),
        ('Dosage Rate (per Bird)', purple_fill),
        ('Total Dosage',           purple_fill),
        ('Feed Name',              orange_fill),
        ('Feed Item ID',           orange_fill),
        ('Feed Rate (per Bird)',   orange_fill),
        ('Total Feed',             orange_fill),
        ('Feed Unit',              orange_fill),
        ('Remarks',                green_fill),
    ]

    for col, (header, fill) in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = fill
        cell.alignment = center

    col_widths = [10, 16, 12, 22, 16, 8, 22, 14, 22, 16, 20, 14, 12, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    for row_num, step in enumerate(program.steps.all(), 2):
        age_display   = f'Day {step.day}' if step.week == 0 else f'Week {step.week}-Day {step.day}'
        medicine_name = material_map.get(step.medicine, step.medicine)
        feed_name     = material_map.get(step.feed, step.feed)
        total_dosage  = float(step.dose_amount) * total_birds if step.dose_amount else ''
        total_feed    = float(step.feed_rate_per_bird) * total_birds if step.feed_rate_per_bird else ''

        row_data = [
            step.cycle,
            age_display,
            total_birds,
            medicine_name,
            step.medicine or '',
            step.dose_unit or '',
            float(step.dose_amount) if step.dose_amount else '',
            total_dosage,
            feed_name,
            step.feed or '',
            float(step.feed_rate_per_bird) if step.feed_rate_per_bird else '',
            total_feed,
            step.feed_unit or '',
            step.remarks or '',
        ]

        for col, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = left

        if row_num % 2 == 0:
            light_fill = PatternFill(fill_type='solid', fgColor='EBF3FB')
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = light_fill

    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{program.name}_{program.type}.xlsx"'
    wb.save(response)
    return response


@login_required
def import_program(request, pk):
    import openpyxl
    from master_data.models import Material

    program = get_object_or_404(Program, pk=pk)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        wb   = openpyxl.load_workbook(file)
        ws   = wb.active

        valid_ids = set(Material.objects.values_list('item_id', flat=True))
        program.steps.all().delete()

        imported = 0
        errors   = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                cycle       = row[0]
                age_display = str(row[1]).strip() if row[1] not in (None, '') else ''
                # row[2] = Population — skip
                # row[3] = Medicine Name — skip
                medicine_id = str(row[4]).strip() if row[4] not in (None, '') else ''
                dose_unit   = str(row[5]).strip() if row[5] not in (None, '') else ''
                dose_amount = row[6]
                # row[7] = Total Dosage — skip
                # row[8] = Feed Name — skip
                feed_id     = str(row[9]).strip() if row[9] not in (None, '') else ''
                feed_rate   = row[10]
                # row[11] = Total Feed — skip
                feed_unit   = str(row[12]).strip() if row[12] not in (None, '') else ''
                remarks     = str(row[13]).strip() if row[13] not in (None, '') else ''

                if medicine_id and medicine_id not in valid_ids:
                    errors.append(f'Row {row_num}: Medicine "{medicine_id}" not found — skipped')
                    continue
                if feed_id and feed_id not in valid_ids:
                    errors.append(f'Row {row_num}: Feed "{feed_id}" not found — skipped')
                    continue

                week = 0
                day  = 1
                if age_display.startswith('Week'):
                    parts = age_display.replace('Week ', '').split('-Day ')
                    week  = int(parts[0].strip())
                    day   = int(parts[1].strip())
                elif age_display.startswith('Day'):
                    day  = int(age_display.replace('Day ', '').strip())
                    week = 0

                ProgramStep.objects.create(
                    program            = program,
                    cycle              = int(cycle) if cycle else 1,
                    week               = week,
                    day                = day,
                    medicine           = medicine_id,
                    dose_amount        = dose_amount if dose_amount not in (None, '') else None,
                    dose_unit          = dose_unit,
                    dose_per           = 'chick',
                    method             = '',
                    remarks            = remarks,
                    feed               = feed_id,
                    feed_rate_per_bird = feed_rate if feed_rate not in (None, '') else None,
                    feed_unit          = feed_unit,
                )
                imported += 1

            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')

        if errors:
            messages.warning(request, f'Imported {imported} steps with {len(errors)} errors: {"; ".join(errors[:3])}')
        else:
            messages.success(request, f'Successfully imported {imported} steps!')

    if program.type == 'Laying':
        return redirect('program_laying')
    return redirect('program_growing')


@login_required
def view_growing(request):
    from flock.models import Flock
    from consumption.models import Consumption
    from erp_config.models import Building
    from master_data.models import Material
    from django.db.models import Sum
    from datetime import timedelta

    programs         = Program.objects.filter(type='Growing').prefetch_related('steps', 'batches')
    material_map     = {m.item_id: m.name for m in Material.objects.all()}
    material_objects = {m.item_id: m for m in Material.objects.all()}

    # build consumption lookup: {item_id: {growing_house: {date: quantity}}}
    consumption_lookup = {}
    consumptions = Consumption.objects.values(
                    'item_id', 'growing_house', 'date_consumed'
                   ).annotate(total=Sum('quantity'))
    for c in consumptions:
        iid   = c['item_id']
        house = c['growing_house']
        date  = str(c['date_consumed'])
        if iid not in consumption_lookup:
            consumption_lookup[iid] = {}
        if house not in consumption_lookup[iid]:
            consumption_lookup[iid][house] = {}
        consumption_lookup[iid][house][date] = float(c['total'])

    program_summaries = []
    for program in programs:
        steps   = list(program.steps.all())
        batches = list(program.batches.filter(flock_type='Growing'))

        batch_summaries = []
        for batch in batches:
            flock = batch.flock
            if not flock:
                continue
            population = batch.population
            house      = batch.flock_house

            # ── Medicine ─────────────────────────────
            med_map = {}
            for step in steps:
                iid = step.medicine
                if not iid or not step.dose_amount:
                    continue
                if iid not in med_map:
                    med_map[iid] = {
                        'name':     material_map.get(iid, iid),
                        'unit':     step.dose_unit,
                        'planned':  0,
                        'consumed': 0,
                        'daily':    [],
                    }
                planned_day  = float(step.dose_amount) * population
                med_map[iid]['planned'] += planned_day
                step_date    = batch.step_date(step.week, step.day)
                date_str     = str(step_date)
                consumed_base = consumption_lookup.get(iid, {}).get(house, {}).get(date_str, 0)
                material      = material_objects.get(iid)

                if material:
                    consumed_converted, can_convert = convert_to_program_unit(
                        consumed_base, step.dose_unit, material)
                else:
                    consumed_converted, can_convert = consumed_base, True

                if can_convert:
                    consumed_day = consumed_converted or 0
                    med_map[iid]['consumed'] += consumed_day
                    remaining = planned_day - consumed_day
                    status    = 'met' if consumed_day >= planned_day else ('short' if consumed_day > 0 else 'none')
                    med_map[iid]['daily'].append({
                        'date': date_str, 'planned': planned_day,
                        'consumed': consumed_day, 'remaining': remaining,
                        'status': status, 'has_date': True, 'unit_ok': True,
                    })
                else:
                    med_map[iid]['daily'].append({
                        'date': date_str, 'planned': planned_day,
                        'consumed': 0, 'remaining': planned_day,
                        'status': 'unit_mismatch', 'has_date': True, 'unit_ok': False,
                    })

            for m in med_map.values():
                m['percentage'] = round(
                    (m['consumed'] / m['planned'] * 100), 1
                ) if m['planned'] > 0 else 0

            # ── Feed ─────────────────────────────────
            feed_map = {}
            for step in steps:
                iid = step.feed
                if not iid or not step.feed_rate_per_bird:
                    continue
                if iid not in feed_map:
                    feed_map[iid] = {
                        'name':     material_map.get(iid, iid),
                        'unit':     step.feed_unit,
                        'planned':  0,
                        'consumed': 0,
                        'daily':    [],
                    }
                planned_day   = float(step.feed_rate_per_bird) * population
                feed_map[iid]['planned'] += planned_day
                step_date     = batch.step_date(step.week, step.day)
                date_str      = str(step_date)
                consumed_base = consumption_lookup.get(iid, {}).get(house, {}).get(date_str, 0)
                material      = material_objects.get(iid)

                if material:
                    consumed_converted, can_convert = convert_to_program_unit(
                        consumed_base, step.feed_unit, material)
                else:
                    consumed_converted, can_convert = consumed_base, True

                if can_convert:
                    consumed_day = consumed_converted or 0
                    feed_map[iid]['consumed'] += consumed_day
                    remaining = planned_day - consumed_day
                    status    = 'met' if consumed_day >= planned_day else ('short' if consumed_day > 0 else 'none')
                    feed_map[iid]['daily'].append({
                        'date': date_str, 'planned': planned_day,
                        'consumed': consumed_day, 'remaining': remaining,
                        'status': status, 'has_date': True, 'unit_ok': True,
                    })
                else:
                    feed_map[iid]['daily'].append({
                        'date': date_str, 'planned': planned_day,
                        'consumed': 0, 'remaining': planned_day,
                        'status': 'unit_mismatch', 'has_date': True, 'unit_ok': False,
                    })

            for f in feed_map.values():
                f['percentage'] = round(
                    (f['consumed'] / f['planned'] * 100), 1
                ) if f['planned'] > 0 else 0

            batch_summaries.append({
                'batch':     batch,
                'medicines': list(med_map.values()),
                'feeds':     list(feed_map.values()),
            })

        program_summaries.append({
            'program':          program,
            'batch_summaries':  batch_summaries,
        })

    return render(request, 'program/view_growing.html', {
        'program_summaries': program_summaries,
    })


@login_required
def view_laying(request):
    from laying_flock.models import LayingFlock
    from consumption.models import Consumption
    from erp_config.models import Building
    from master_data.models import Material
    from django.db.models import Sum
    from datetime import timedelta

    programs         = Program.objects.filter(type='Laying').prefetch_related('steps', 'batches')
    material_map     = {m.item_id: m.name for m in Material.objects.all()}
    material_objects = {m.item_id: m for m in Material.objects.all()}

    consumption_lookup = {}
    consumptions = Consumption.objects.values(
                    'item_id', 'growing_house', 'date_consumed'
                   ).annotate(total=Sum('quantity'))
    for c in consumptions:
        iid   = c['item_id']
        house = c['growing_house']
        date  = str(c['date_consumed'])
        if iid not in consumption_lookup:
            consumption_lookup[iid] = {}
        if house not in consumption_lookup[iid]:
            consumption_lookup[iid][house] = {}
        consumption_lookup[iid][house][date] = float(c['total'])

    program_summaries = []
    for program in programs:
        steps   = list(program.steps.all())
        batches = list(program.batches.filter(flock_type='Laying'))

        batch_summaries = []
        for batch in batches:
            flock = batch.flock
            if not flock:
                continue
            population = batch.population
            house      = batch.flock_house

            med_map = {}
            for step in steps:
                iid = step.medicine
                if not iid or not step.dose_amount:
                    continue
                if iid not in med_map:
                    med_map[iid] = {
                        'name':     material_map.get(iid, iid),
                        'unit':     step.dose_unit,
                        'planned':  0,
                        'consumed': 0,
                        'daily':    [],
                    }
                planned_day   = float(step.dose_amount) * population
                med_map[iid]['planned'] += planned_day
                step_date     = batch.step_date(step.week, step.day)
                date_str      = str(step_date)
                consumed_base = consumption_lookup.get(iid, {}).get(house, {}).get(date_str, 0)
                material      = material_objects.get(iid)

                if material:
                    consumed_converted, can_convert = convert_to_program_unit(
                        consumed_base, step.dose_unit, material)
                else:
                    consumed_converted, can_convert = consumed_base, True

                if can_convert:
                    consumed_day = consumed_converted or 0
                    med_map[iid]['consumed'] += consumed_day
                    remaining = planned_day - consumed_day
                    status    = 'met' if consumed_day >= planned_day else ('short' if consumed_day > 0 else 'none')
                    med_map[iid]['daily'].append({
                        'date': date_str, 'planned': planned_day,
                        'consumed': consumed_day, 'remaining': remaining,
                        'status': status, 'has_date': True, 'unit_ok': True,
                    })
                else:
                    med_map[iid]['daily'].append({
                        'date': date_str, 'planned': planned_day,
                        'consumed': 0, 'remaining': planned_day,
                        'status': 'unit_mismatch', 'has_date': True, 'unit_ok': False,
                    })

            for m in med_map.values():
                m['percentage'] = round(
                    (m['consumed'] / m['planned'] * 100), 1
                ) if m['planned'] > 0 else 0

            feed_map = {}
            for step in steps:
                iid = step.feed
                if not iid or not step.feed_rate_per_bird:
                    continue
                if iid not in feed_map:
                    feed_map[iid] = {
                        'name':     material_map.get(iid, iid),
                        'unit':     step.feed_unit,
                        'planned':  0,
                        'consumed': 0,
                        'daily':    [],
                    }
                planned_day   = float(step.feed_rate_per_bird) * population
                feed_map[iid]['planned'] += planned_day
                step_date     = batch.step_date(step.week, step.day)
                date_str      = str(step_date)
                consumed_base = consumption_lookup.get(iid, {}).get(house, {}).get(date_str, 0)
                material      = material_objects.get(iid)

                if material:
                    consumed_converted, can_convert = convert_to_program_unit(
                        consumed_base, step.feed_unit, material)
                else:
                    consumed_converted, can_convert = consumed_base, True

                if can_convert:
                    consumed_day = consumed_converted or 0
                    feed_map[iid]['consumed'] += consumed_day
                    remaining = planned_day - consumed_day
                    status    = 'met' if consumed_day >= planned_day else ('short' if consumed_day > 0 else 'none')
                    feed_map[iid]['daily'].append({
                        'date': date_str, 'planned': planned_day,
                        'consumed': consumed_day, 'remaining': remaining,
                        'status': status, 'has_date': True, 'unit_ok': True,
                    })
                else:
                    feed_map[iid]['daily'].append({
                        'date': date_str, 'planned': planned_day,
                        'consumed': 0, 'remaining': planned_day,
                        'status': 'unit_mismatch', 'has_date': True, 'unit_ok': False,
                    })

            for f in feed_map.values():
                f['percentage'] = round(
                    (f['consumed'] / f['planned'] * 100), 1
                ) if f['planned'] > 0 else 0

            batch_summaries.append({
                'batch':     batch,
                'medicines': list(med_map.values()),
                'feeds':     list(feed_map.values()),
            })

        program_summaries.append({
            'program':         program,
            'batch_summaries': batch_summaries,
        })

    return render(request, 'program/view_laying.html', {
        'program_summaries': program_summaries,
    })